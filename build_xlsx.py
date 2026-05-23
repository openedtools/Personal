"""Build Desert Turkey 2026 trip planner workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
import datetime, urllib.request, json, ssl

# ── Palette ──────────────────────────────────────────────────
CREAM         = "FFF5EED8"
SURFACE       = "FFEDE4CC"
SURFACE2      = "FFE5D8B8"
BORDER_CLR    = "FFC8B888"
TERRACOTTA    = "FFC4622D"
TEXT          = "FF2A1A08"
MUTED         = "FF7A6040"

# Camper palette (cycled by row mod 7) — mirrors the website
CAMPER_COLORS = ["FFC4622D", "FF2D6BC4", "FF6BC42D", "FFC42D8A",
                 "FF2DC4B0", "FFC4A02D", "FF8A2DC4"]

THIN = Side(border_style="thin", color=BORDER_CLR)

# ── Helpers ──────────────────────────────────────────────────
def header_style(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=TERRACOTTA)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def title_style(cell, size=22):
    cell.font = Font(name="Calibri", size=size, bold=True, color=TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def stat_label(cell):
    cell.font = Font(name="Calibri", size=10, italic=True, color=MUTED)
    cell.alignment = Alignment(horizontal="left")

def stat_value(cell):
    cell.font = Font(name="Calibri", size=18, bold=True, color=TERRACOTTA)
    cell.alignment = Alignment(horizontal="left")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def cream_background(ws, rows=80, cols=12):
    fill = PatternFill("solid", fgColor=CREAM)
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).fill = fill

def add_table(ws, name, ref, style="TableStyleLight15"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    return tbl

# ── Fetch current Firestore data ─────────────────────────────
def fetch_firestore():
    url = ("https://firestore.googleapis.com/v1/projects/joshua-tree-26/"
           "databases/(default)/documents/trips/thanksgiving-2026")
    try:
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(url, timeout=10, context=ctx) as r:
            doc = json.load(r)
    except Exception as e:
        print(f"(could not fetch Firestore: {e})")
        return {}
    return unwrap(doc.get("fields", {}))

def unwrap(node):
    if not isinstance(node, dict):
        return node
    if "stringValue"  in node: return node["stringValue"]
    if "integerValue" in node: return int(node["integerValue"])
    if "doubleValue"  in node: return float(node["doubleValue"])
    if "booleanValue" in node: return node["booleanValue"]
    if "nullValue"    in node: return None
    if "arrayValue"   in node:
        return [unwrap(v) for v in node["arrayValue"].get("values", [])]
    if "mapValue"     in node:
        return {k: unwrap(v) for k, v in node["mapValue"].get("fields", {}).items()}
    return {k: unwrap(v) for k, v in node.items()}

data = fetch_firestore()
campers      = data.get("campers", [])
reservations = data.get("siteReservations", [])
potluck      = data.get("potluck", [])
tshirts      = data.get("tshirts", [])
itinerary    = data.get("itinerary", [])

print(f"Fetched: {len(campers)} campers, {len(reservations)} reservations, "
      f"{len(potluck)} potluck, {len(tshirts)} t-shirts, {len(itinerary)} itinerary")

# ── Build workbook ───────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

TRIP_DAYS = [datetime.date(2026, 11, d) for d in range(22, 29)]  # 22..28

# ═══════════════════════════════════════════════
# Welcome tab (dashboard)
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Welcome")
cream_background(ws, rows=40, cols=8)
set_col_widths(ws, [22, 22, 22, 22, 22, 22])
ws.sheet_view.showGridLines = False

ws["A1"] = "Desert Turkey 2026"
title_style(ws["A1"], size=28)
ws.row_dimensions[1].height = 38

ws["A2"] = "Jumbo Rocks Campground · Joshua Tree NP · Nov 22–28, 2026"
ws["A2"].font = Font(name="Calibri", size=12, italic=True, color=MUTED)

ws["A4"] = "Countdown"
stat_label(ws["A4"])
ws["A5"] = '=MAX(0, DATE(2026,11,22)-TODAY()) & " days to go"'
ws["A5"].font = Font(name="Calibri", size=20, bold=True, color=TERRACOTTA)

ws["A7"] = "Campers";           stat_label(ws["A7"])
ws["A8"] = "=COUNTA(Campers[Name])"; stat_value(ws["A8"])

ws["B7"] = "People (adults + kids)"; stat_label(ws["B7"])
ws["B8"] = "=SUM(Campers[Adults])+SUM(Campers[Kids])"; stat_value(ws["B8"])

ws["C7"] = "Reservations"; stat_label(ws["C7"])
ws["C8"] = "=COUNTA(Reservations[Holder])"; stat_value(ws["C8"])

ws["D7"] = "Sites assigned"; stat_label(ws["D7"])
ws["D8"] = '=COUNTIFS(SiteAssignments[Used By], "<>")'; stat_value(ws["D8"])

ws["E7"] = "Sites open"; stat_label(ws["E7"])
ws["E8"] = '=COUNTIFS(SiteAssignments[Used By], "")'; stat_value(ws["E8"])

ws["A11"] = "Tabs"
title_style(ws["A11"], size=14)
tab_notes = [
    ("Campers",         "Add yourself: name, party size, setup, arrival/departure"),
    ("Timeline",        "Gantt view auto-built from Campers"),
    ("Reservations",    "Who holds each Recreation.gov booking"),
    ("Site Assignments","Per-site occupant tracker"),
    ("Itinerary",       "Editable trip plan (currently TBD)"),
    ("Potluck",         "Sign up for what you're bringing"),
    ("T-Shirts",        "Order tally by size"),
    ("Gear",            "Suggested packing list"),
    ("Trails",          "Nearby hikes with NPS/AllTrails links"),
    ("Links & Info",    "Reservation, weather, dark sky, hospital, etc."),
]
for i, (tab, desc) in enumerate(tab_notes, start=12):
    ws.cell(row=i, column=1, value=tab).font = Font(bold=True, color=TERRACOTTA)
    ws.cell(row=i, column=2, value=desc).font = Font(color=TEXT)

# ═══════════════════════════════════════════════
# Campers tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Campers")
cream_background(ws, rows=40, cols=8)
set_col_widths(ws, [26, 10, 8, 16, 14, 14, 36])
ws.sheet_view.showGridLines = False

headers = ["Name", "Adults", "Kids", "Setup", "Arrival", "Departure", "Note"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

# Body rows — pre-fill from Firestore or pad with blanks
PAD_CAMPERS = max(20, len(campers) + 5)
for i in range(PAD_CAMPERS):
    r = i + 2
    c = campers[i] if i < len(campers) else None
    if c:
        ws.cell(row=r, column=1, value=c.get("name", ""))
        ws.cell(row=r, column=2, value=int(c.get("adults", 1) or 1))
        ws.cell(row=r, column=3, value=int(c.get("kids",   0) or 0))
        ws.cell(row=r, column=4, value=c.get("setup", "Tent"))
        if c.get("arrival"):
            ws.cell(row=r, column=5,
                    value=datetime.datetime.strptime(c["arrival"], "%Y-%m-%d").date()).number_format = "ddd m/d"
        if c.get("departure"):
            ws.cell(row=r, column=6,
                    value=datetime.datetime.strptime(c["departure"], "%Y-%m-%d").date()).number_format = "ddd m/d"
        ws.cell(row=r, column=7, value=c.get("note", ""))
    # Ensure date format for empty cells too so date pickers know
    if not ws.cell(row=r, column=5).number_format.startswith("ddd"):
        ws.cell(row=r, column=5).number_format = "ddd m/d"
        ws.cell(row=r, column=6).number_format = "ddd m/d"

last_row = 1 + PAD_CAMPERS
add_table(ws, "Campers", f"A1:G{last_row}")

# Data validation: Setup dropdown
dv_setup = DataValidation(type="list",
    formula1='"Tent,Van,RV,Car camping,Cabin"', allow_blank=True)
dv_setup.add(f"D2:D{last_row}")
ws.add_data_validation(dv_setup)

# Data validation: dates between 2026-11-22 and 2026-11-28
dv_arr = DataValidation(type="date", operator="between",
    formula1=datetime.date(2026,11,22),
    formula2=datetime.date(2026,11,28),
    allow_blank=True,
    error="Trip runs Nov 22–28, 2026.", errorTitle="Date out of range")
dv_arr.add(f"E2:F{last_row}")
ws.add_data_validation(dv_arr)

# Conditional formatting: row tint by camper index (mod 7) — light bar on Name col only
for i, col in enumerate(CAMPER_COLORS):
    light = "FF" + ''.join(
        format(min(255, int(col[2*j+2:2*j+4], 16) + 90), '02X') for j in range(3)
    )
    rule = FormulaRule(
        formula=[f'AND($A2<>"", MOD(ROW()-2,7)={i})'],
        fill=PatternFill("solid", fgColor=light))
    ws.conditional_formatting.add(f"A2:G{last_row}", rule)

ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# Timeline tab (Gantt)
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Timeline")
cream_background(ws, rows=40, cols=10)
set_col_widths(ws, [26] + [12]*7)
ws.sheet_view.showGridLines = False

ws["A1"] = "Camper"; header_style(ws["A1"])
for j, d in enumerate(TRIP_DAYS):
    cell = ws.cell(row=1, column=2+j, value=d); header_style(cell)
    cell.number_format = "ddd\nm/d"
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 36
ws.freeze_panes = "B2"

# Body: 20 camper rows pulled from Campers table by position
for i in range(20):
    r = i + 2
    ws.cell(row=r, column=1, value=f'=IFERROR(INDEX(Campers[Name],{i+1}),"")')

# Conditional formatting per camper row (mod 7) — colored bar when present that day
# Range: B2:H{last}
last_row_t = 21
for i, col in enumerate(CAMPER_COLORS):
    rule = FormulaRule(
        formula=[(
            f'AND(MOD(ROW()-2,7)={i},'
            f'INDEX(Campers[Name],ROW()-1)<>"",'
            f'B$1>=INDEX(Campers[Arrival],ROW()-1),'
            f'B$1<=INDEX(Campers[Departure],ROW()-1))'
        )],
        fill=PatternFill("solid", fgColor=col))
    ws.conditional_formatting.add(f"B2:H{last_row_t}", rule)

# Subtle empty-cell border so the grid reads as a timeline even when blank
for r in range(2, last_row_t + 1):
    for c in range(2, 9):
        cell = ws.cell(row=r, column=c)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal="center", vertical="center")

ws["A23"] = "How this works"
ws["A23"].font = Font(bold=True, color=TEXT)
ws["A24"] = ("Rows auto-populate from the Campers tab in order. Each cell fills "
             "when that day falls between the camper's Arrival and Departure dates.")
ws["A24"].font = Font(italic=True, color=MUTED, size=10)
ws["A24"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A24:H25")

# ═══════════════════════════════════════════════
# Reservations tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Reservations")
cream_background(ws, rows=40, cols=8)
set_col_widths(ws, [16, 24, 24, 14, 14, 36])
ws.sheet_view.showGridLines = False

headers = ["ID", "Holder", "Sites (CSV)", "Arrival", "Departure", "Note"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

PAD_RES = max(10, len(reservations) + 3)
for i in range(PAD_RES):
    r = i + 2
    res = reservations[i] if i < len(reservations) else None
    if res:
        ws.cell(row=r, column=1, value=res.get("id", ""))
        ws.cell(row=r, column=2, value=res.get("reservedBy", ""))
        sites_csv = ", ".join(s.get("siteNum", "") for s in res.get("sites", []) or [])
        ws.cell(row=r, column=3, value=sites_csv)
        if res.get("arrival"):
            c = ws.cell(row=r, column=4, value=datetime.datetime.strptime(res["arrival"], "%Y-%m-%d").date())
            c.number_format = "ddd m/d"
        if res.get("departure"):
            c = ws.cell(row=r, column=5, value=datetime.datetime.strptime(res["departure"], "%Y-%m-%d").date())
            c.number_format = "ddd m/d"
        ws.cell(row=r, column=6, value=res.get("note", ""))
    ws.cell(row=r, column=4).number_format = "ddd m/d"
    ws.cell(row=r, column=5).number_format = "ddd m/d"

add_table(ws, "Reservations", f"A1:F{1+PAD_RES}")
ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# Site Assignments tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Site Assignments")
cream_background(ws, rows=50, cols=6)
set_col_widths(ws, [22, 12, 24])
ws.sheet_view.showGridLines = False

headers = ["Reservation ID", "Site #", "Used By"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

# Flatten reservation sites
flat_sites = []
for res in reservations:
    for s in (res.get("sites") or []):
        flat_sites.append((res.get("id", ""), s.get("siteNum", ""), s.get("usedBy", "")))

PAD_SITES = max(30, len(flat_sites) + 10)
for i in range(PAD_SITES):
    r = i + 2
    if i < len(flat_sites):
        rid, snum, used = flat_sites[i]
        ws.cell(row=r, column=1, value=rid)
        ws.cell(row=r, column=2, value=snum)
        ws.cell(row=r, column=3, value=used)

add_table(ws, "SiteAssignments", f"A1:C{1+PAD_SITES}")

# CF: open sites shown muted-italic
open_font  = Font(italic=True, color=MUTED)
open_fill  = PatternFill("solid", fgColor=SURFACE)
rule_open = FormulaRule(formula=['AND($A2<>"", $C2="")'], font=open_font, fill=open_fill)
ws.conditional_formatting.add(f"A2:C{1+PAD_SITES}", rule_open)

ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# Itinerary tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Itinerary")
cream_background(ws, rows=40, cols=6)
set_col_widths(ws, [10, 16, 28, 60])
ws.sheet_view.showGridLines = False

headers = ["Day", "Date", "Title", "Activities (one per line — Alt+Enter)"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

PAD_IT = max(10, len(itinerary) + 5)
for i in range(PAD_IT):
    r = i + 2
    d = itinerary[i] if i < len(itinerary) else None
    if d:
        ws.cell(row=r, column=1, value=d.get("day", ""))
        ws.cell(row=r, column=2, value=d.get("date", ""))
        ws.cell(row=r, column=3, value=d.get("title", ""))
        ws.cell(row=r, column=4, value="\n".join(d.get("activities") or []))
    ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 60

add_table(ws, "Itinerary", f"A1:D{1+PAD_IT}")
ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# Potluck tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Potluck")
cream_background(ws, rows=40, cols=6)
set_col_widths(ws, [22, 30, 16, 36])
ws.sheet_view.showGridLines = False

headers = ["Name", "Bringing", "Category", "Note"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

PAD_PL = max(20, len(potluck) + 5)
for i in range(PAD_PL):
    r = i + 2
    p = potluck[i] if i < len(potluck) else None
    if p:
        ws.cell(row=r, column=1, value=p.get("name", ""))
        ws.cell(row=r, column=2, value=p.get("dish", ""))
        ws.cell(row=r, column=3, value=(p.get("category") or "").title())
        ws.cell(row=r, column=4, value=p.get("note", ""))

add_table(ws, "Potluck", f"A1:D{1+PAD_PL}")

# Category dropdown
dv_cat = DataValidation(type="list",
    formula1='"Main,Side,Appetizer,Dessert,Drinks,Snacks,Other"', allow_blank=True)
dv_cat.add(f"C2:C{1+PAD_PL}")
ws.add_data_validation(dv_cat)

# Category color CF (cell-level, on Category column)
cat_colors = {
    "Main":      "FFFCE8D8",
    "Side":      "FFFEF0C8",
    "Appetizer": "FFE8F0D8",
    "Dessert":   "FFF8E0F0",
    "Drinks":    "FFD8ECF8",
    "Snacks":    "FFE8E8F8",
    "Other":     "FFE5D8B8",
}
for cat, col in cat_colors.items():
    rule = FormulaRule(formula=[f'$C2="{cat}"'],
                       fill=PatternFill("solid", fgColor=col))
    ws.conditional_formatting.add(f"C2:C{1+PAD_PL}", rule)

# Summary panel
ws["F1"] = "Count by category"; header_style(ws["F1"])
ws.column_dimensions["F"].width = 22
ws.column_dimensions["G"].width = 10
for i, cat in enumerate(cat_colors.keys(), start=2):
    ws.cell(row=i, column=6, value=cat).font = Font(color=TEXT)
    ws.cell(row=i, column=7, value=f'=COUNTIF(Potluck[Category], "{cat}")').font = \
        Font(bold=True, color=TERRACOTTA)
ws.cell(row=2+len(cat_colors), column=6, value="Total").font = Font(bold=True, color=TEXT)
ws.cell(row=2+len(cat_colors), column=7,
        value="=COUNTA(Potluck[Name])").font = Font(bold=True, color=TERRACOTTA)

ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# T-Shirts tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("T-Shirts")
cream_background(ws, rows=40, cols=8)
set_col_widths(ws, [22, 10, 8, 30])
ws.sheet_view.showGridLines = False

headers = ["Name", "Size", "Qty", "Note"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

PAD_TS = max(20, len(tshirts) + 5)
for i in range(PAD_TS):
    r = i + 2
    t = tshirts[i] if i < len(tshirts) else None
    if t:
        ws.cell(row=r, column=1, value=t.get("name", ""))
        ws.cell(row=r, column=2, value=t.get("size", ""))
        ws.cell(row=r, column=3, value=int(t.get("qty", 1) or 1))
        ws.cell(row=r, column=4, value=t.get("note", ""))

add_table(ws, "TShirts", f"A1:D{1+PAD_TS}")

dv_size = DataValidation(type="list",
    formula1='"XS,S,M,L,XL,2XL"', allow_blank=True)
dv_size.add(f"B2:B{1+PAD_TS}")
ws.add_data_validation(dv_size)

# Summary
ws["F1"] = "Tally by size"; header_style(ws["F1"])
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 10
sizes = ["XS","S","M","L","XL","2XL"]
for i, s in enumerate(sizes, start=2):
    ws.cell(row=i, column=6, value=s).font = Font(color=TEXT)
    ws.cell(row=i, column=7, value=f'=SUMIF(TShirts[Size], "{s}", TShirts[Qty])').font = \
        Font(bold=True, color=TERRACOTTA)
ws.cell(row=2+len(sizes), column=6, value="Total").font = Font(bold=True, color=TEXT)
ws.cell(row=2+len(sizes), column=7, value="=SUM(TShirts[Qty])").font = Font(bold=True, color=TERRACOTTA)

ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# Gear tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Gear")
cream_background(ws, rows=60, cols=4)
set_col_widths(ws, [36, 16])
ws.sheet_view.showGridLines = False

headers = ["Item", "Category"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

GEAR = [
    ("Tent (4-person)", "Shelter"),
    ("Sleeping bag (20°F rated)", "Shelter"),
    ("Sleeping pad or air mattress", "Shelter"),
    ("Camp chairs", "Shelter"),
    ("Folding table", "Shelter"),
    ("Camp stove + fuel", "Cooking"),
    ("Cooler with ice", "Cooking"),
    ("Cast iron skillet", "Cooking"),
    ("Cooking pot", "Cooking"),
    ("Plates, cups, utensils", "Cooking"),
    ("Turkey (or protein main)", "Food"),
    ("Stuffing, mashed potatoes, sides", "Food"),
    ("Firewood (2 bundles/night)", "Food"),
    ("Coffee + coffee maker", "Food"),
    ("Drinks & cooler drinks", "Food"),
    ("S'mores kit", "Food"),
    ("First aid kit", "Safety"),
    ("Headlamps (+ extra batteries)", "Tools"),
    ("Lantern", "Tools"),
    ("Firestarter + lighter", "Tools"),
    ("Sunscreen + lip balm", "Safety"),
    ("Cards / board games", "Fun"),
    ("Guitar (if applicable)", "Fun"),
    ("Star map / astronomy app", "Fun"),
]
for i, (item, cat) in enumerate(GEAR):
    ws.cell(row=i+2, column=1, value=item)
    ws.cell(row=i+2, column=2, value=cat)

add_table(ws, "Gear", f"A1:B{1+len(GEAR)}")
ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# Trails tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Trails")
cream_background(ws, rows=20, cols=8)
set_col_widths(ws, [28, 12, 12, 12, 14, 50, 14, 14])
ws.sheet_view.showGridLines = False

headers = ["Name", "Difficulty", "Distance", "Time", "Elevation",
           "Description", "NPS link", "AllTrails link"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

TRAILS = [
    ("Skull Rock Nature Trail", "Easy", "1.7 mi", "~1 hr", "minimal",
     "Starts at Jumbo Rocks. Wind through boulder formations to the iconic skull-shaped rock.",
     "https://www.nps.gov/places/skull-rock-nature-trail.htm",
     "https://www.alltrails.com/trail/us/california/skull-rock-trail"),
    ("Ryan Mountain Trail", "Moderate", "3.0 mi", "2–3 hrs", "+1,070 ft",
     "Best panoramic summit in the park. Mt. San Jacinto, San Gorgonio, and the Salton Sea on a clear day.",
     "https://www.nps.gov/places/ryan-mountain.htm",
     "https://www.alltrails.com/trail/us/california/ryan-mountain-trail"),
    ("Arch Rock Nature Trail", "Easy", "1.3 mi", "~45 min", "minimal",
     "Short loop through boulders to a massive natural arch.",
     "https://www.nps.gov/places/arch-rock-nature-trail.htm",
     "https://www.alltrails.com/trail/us/california/arch-rock-trail"),
    ("Hidden Valley Nature Trail", "Easy", "1.0 mi", "~45 min", "minimal",
     "Classic loop through a boulder-ringed valley once used by cattle rustlers.",
     "https://www.nps.gov/places/hidden-valley-nature-trail.htm",
     "https://www.alltrails.com/trail/us/california/hidden-valley-nature-trail"),
    ("Cholla Cactus Garden Loop", "Easy", "0.25 mi", "20 min", "flat",
     "Alien landscape of dense teddy bear cholla. Watch out — they grab you.",
     "https://www.nps.gov/places/cholla-cactus-garden.htm",
     "https://www.alltrails.com/trail/us/california/cholla-cactus-garden-nature-trail"),
    ("Split Rock Loop", "Easy", "2.0 mi", "~1.5 hrs", "minimal",
     "Huge boulder split cleanly in two by erosion. Loop through desert scenery.",
     "https://www.nps.gov/places/split-rock.htm",
     "https://www.alltrails.com/trail/us/california/split-rock-loop-trail"),
    ("Keys View", "Easy", "0.25 mi (paved)", "15 min", "drive-up",
     "Stunning views across Coachella Valley, San Andreas Fault, Salton Sea. Best at sunset.",
     "https://www.nps.gov/places/keys-view.htm",
     "https://www.alltrails.com/trail/us/california/keys-view"),
    ("Barker Dam Nature Trail", "Easy", "1.3 mi", "~1 hr", "minimal",
     "Historic stone dam with a natural water tank. Bighorn sheep frequent.",
     "https://www.nps.gov/places/barker-dam-nature-trail.htm",
     "https://www.alltrails.com/trail/us/california/barker-dam-trail"),
]
for i, t in enumerate(TRAILS):
    r = i + 2
    ws.cell(row=r, column=1, value=t[0])
    ws.cell(row=r, column=2, value=t[1])
    ws.cell(row=r, column=3, value=t[2])
    ws.cell(row=r, column=4, value=t[3])
    ws.cell(row=r, column=5, value=t[4])
    ws.cell(row=r, column=6, value=t[5]).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=7, value=f'=HYPERLINK("{t[6]}","NPS ↗")').font = Font(color=TERRACOTTA, underline="single")
    ws.cell(row=r, column=8, value=f'=HYPERLINK("{t[7]}","AllTrails ↗")').font = Font(color=TERRACOTTA, underline="single")
    ws.row_dimensions[r].height = 42

add_table(ws, "Trails", f"A1:H{1+len(TRAILS)}")

# Difficulty CF
diff_colors = {"Easy": "FFD8F0D0", "Moderate": "FFF8E8B8", "Hard": "FFF8D8C8"}
for d, col in diff_colors.items():
    rule = FormulaRule(formula=[f'$B2="{d}"'],
                       fill=PatternFill("solid", fgColor=col))
    ws.conditional_formatting.add(f"B2:B{1+len(TRAILS)}", rule)

ws.freeze_panes = "A2"

# ═══════════════════════════════════════════════
# Links & Info tab
# ═══════════════════════════════════════════════
ws = wb.create_sheet("Links & Info")
cream_background(ws, rows=40, cols=4)
set_col_widths(ws, [28, 60, 30])
ws.sheet_view.showGridLines = False

headers = ["Topic", "Why you'd open it", "Link"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=i, value=h); header_style(cell)
ws.row_dimensions[1].height = 24

LINKS = [
    ("Recreation.gov · Jumbo Rocks",
     "Book / view the campground reservation (campground #272299).",
     "https://www.recreation.gov/camping/campgrounds/272299"),
    ("NPS · Jumbo Rocks page",
     "Official info on the Jumbo Rocks campground (amenities, rules).",
     "https://www.nps.gov/jotr/planyourvisit/jumborocks.htm"),
    ("NPS · Joshua Tree NP",
     "Park homepage — alerts, hours, fees.",
     "https://www.nps.gov/jotr/index.htm"),
    ("NPS · Park Maps",
     "Downloadable park maps for offline use (no cell in the park).",
     "https://www.nps.gov/jotr/planyourvisit/maps.htm"),
    ("NPS · Directions",
     "How to get to each entrance and Jumbo Rocks.",
     "https://www.nps.gov/jotr/planyourvisit/directions.htm"),
    ("NPS · Weather & Climate",
     "What to expect for November temps and wind.",
     "https://www.nps.gov/jotr/planyourvisit/weather.htm"),
    ("NWS · Joshua Tree forecast",
     "National Weather Service forecast for the park.",
     "https://forecast.weather.gov/MapClick.php?lat=33.8734&lon=-115.901"),
    ("NPS · Night Sky",
     "Stargazing tips for the International Dark Sky Park.",
     "https://www.nps.gov/jotr/learn/nature/nightsky.htm"),
    ("DarkSky · Joshua Tree",
     "Official Dark Sky Place listing.",
     "https://darksky.org/places/joshua-tree-national-park/"),
    ("NPS · Animals",
     "Wildlife you might see — and what to do about it.",
     "https://www.nps.gov/jotr/learn/nature/animals.htm"),
    ("NPS · Drinking water",
     "Where to fill up (Jumbo Rocks has NO potable water).",
     "https://www.nps.gov/jotr/planyourvisit/water.htm"),
    ("NPS · Campfires & wood",
     "Fire rules and where to buy firewood.",
     "https://www.nps.gov/jotr/planyourvisit/campfires.htm"),
    ("Visit Twentynine Palms",
     "Nearest town east — gas, groceries, food.",
     "https://visit29.org/"),
    ("Visit Joshua Tree",
     "Town northwest — coffee, quirky shops.",
     "https://visitjoshua.com/"),
    ("Desert Regional Medical Center",
     "Closest ER (Palm Springs, ~45 min).",
     "https://www.desertcareresources.com/dch/locations/desert-regional-medical-center"),
    ("Drive to Jumbo Rocks (Google Maps)",
     "Pin the campground in Maps.",
     "https://www.google.com/maps/place/Jumbo+Rocks+Campground/@34.0108,-116.0503,15z"),
]
for i, (topic, desc, url) in enumerate(LINKS):
    r = i + 2
    ws.cell(row=r, column=1, value=topic).font = Font(bold=True, color=TEXT)
    ws.cell(row=r, column=2, value=desc).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=2).font = Font(color=MUTED)
    ws.cell(row=r, column=3, value=f'=HYPERLINK("{url}","Open ↗")').font = \
        Font(color=TERRACOTTA, underline="single")
    ws.row_dimensions[r].height = 28

add_table(ws, "LinksInfo", f"A1:C{1+len(LINKS)}")
ws.freeze_panes = "A2"

# Reorder tabs
wb._sheets = [wb["Welcome"], wb["Campers"], wb["Timeline"], wb["Reservations"],
              wb["Site Assignments"], wb["Itinerary"], wb["Potluck"], wb["T-Shirts"],
              wb["Gear"], wb["Trails"], wb["Links & Info"]]

# Save
out = "/home/user/Personal/Desert_Turkey_2026_Trip_Planner.xlsx"
wb.save(out)
print(f"Saved: {out}")
